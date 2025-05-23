# excel_processing.py
import openpyxl
from openpyxl import Workbook # Aunque no la usas directamente para crear, es buena práctica si se extiende
from datetime import datetime, timedelta
import os

# Importar configuraciones
import Modulos.config as config

def calculate_arbeitszeit(von_str, bis_str, pause_min_str):
    """
    Calculates the working time in hours.
    Returns float (hours) or None if calculation is not possible.
    """
    try:
        pause_min = int(pause_min_str)
    except (ValueError, TypeError):
        pause_min = 0  # Default to 0 if pause is not a valid number or None

    try:
        if not all(isinstance(s, str) for s in [von_str, bis_str]) or \
                not (":" in von_str and ":" in bis_str) or \
                von_str.lower() in ["unbekannt", "hh:mm", ""] or \
                bis_str.lower() in ["unbekannt", "hh:mm", ""]:
            return None

        time_format = "%H:%M"
        von_dt_base = datetime.strptime(von_str, time_format)
        bis_dt_base = datetime.strptime(bis_str, time_format)
        arbitrary_date = datetime.now().date() # Date doesn't matter, only time part
        von_datetime = datetime.combine(arbitrary_date, von_dt_base.time())
        bis_datetime = datetime.combine(arbitrary_date, bis_dt_base.time())

        if bis_datetime < von_datetime: # Handle overnight work
            bis_datetime += timedelta(days=1)

        duration_total = bis_datetime - von_datetime
        arbeitszeit_delta = duration_total - timedelta(minutes=pause_min)
        arbeitszeit_stunden = arbeitszeit_delta.total_seconds() / 3600.0
        return max(0.0, round(arbeitszeit_stunden, 2)) # Ensure non-negative
    except ValueError:
        return None
    except Exception: # Catch any other unexpected error during calculation
        return None


def fill_excel_bautagebuch(bautagebuch_data):
    """
    Fills the Bautagebuch Excel template with data from the JSON object.
    Uses template_path and output_dir from config.
    """
    template_path = config.EXCEL_TEMPLATE_PATH
    output_dir = config.FILLED_EXCEL_DIR

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except FileNotFoundError:
        print(f"❌ Error: Template file not found at {template_path}")
        return None
    except Exception as e:
        print(f"❌ Error loading Excel template: {e}")
        return None

    # --- Fill simple fields ---
    ws['B3'] = bautagebuch_data.get("Baustelle", "Unbekannt")
    ws['B7'] = bautagebuch_data.get("Auftraggeber_Bauleiter", "Unbekannt")
    ws['B8'] = bautagebuch_data.get("Bauueberwachung_Verantwortlicher", "Unbekannt")
    ws['E3'] = bautagebuch_data.get("Datum", config.get_current_date_str()) # Default to current if not in data
    ws['E4'] = bautagebuch_data.get("Auftrag", "Unbekannt")
    ws['B11'] = bautagebuch_data.get("Wetter", "Unbekannt")
    ws['E11'] = bautagebuch_data.get("Temperatur", "None")
    ws['H11'] = bautagebuch_data.get("Wind", "Unbekannt")

    # --- Fill Personal Data ---
    personal_list = bautagebuch_data.get("Personal", [])
    start_row_personal = 15
    max_personal_entries = 7

    for i in range(max_personal_entries):
        current_row = start_row_personal + i
        if i < len(personal_list):
            person = personal_list[i]
            ws[f'A{current_row}'] = person.get("Baustellenpersonal_Typ", "Unbekannt")
            ws[f'B{current_row}'] = person.get("Name", "Unbekannt")
            von_str = person.get("von", "HH:MM") # Default to placeholder if missing
            bis_str = person.get("bis", "HH:MM") # Default to placeholder if missing
            pause_str = person.get("Pause_Minuten", "0")

            ws[f'C{current_row}'] = von_str
            ws[f'D{current_row}'] = bis_str
            try:
                ws[f'E{current_row}'] = int(pause_str)
            except ValueError:
                ws[f'E{current_row}'] = 0 # Default pause to 0 if invalid

            arbeitszeit_json_str = person.get("Arbeitszeit_Stunden", "0.0")
            final_arbeitszeit = 0.0
            try: # Try to use valid Arbeitszeit_Stunden from JSON if > 0
                json_val_as_float = float(arbeitszeit_json_str)
                if json_val_as_float > 0:
                    final_arbeitszeit = round(json_val_as_float, 2)
                else: # If 0.0 or invalid, trigger calculation
                    raise ValueError("Value is 0.0 or non-positive, attempt calculation")
            except (ValueError, TypeError):
                calculated_zeit = calculate_arbeitszeit(von_str, bis_str, pause_str)
                if calculated_zeit is not None:
                    final_arbeitszeit = calculated_zeit
                else: # Calculation failed, try original JSON value or default to 0.0
                    try:
                        final_arbeitszeit = round(float(arbeitszeit_json_str), 2)
                    except (ValueError, TypeError):
                        final_arbeitszeit = 0.0
            ws[f'F{current_row}'] = final_arbeitszeit
        else: # Clear rows if fewer entries than max
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                 ws[f'{col_letter}{current_row}'] = ""

    # --- Fill other text fields ---
    ws['C24'] = bautagebuch_data.get("Ausgefuehrte_Arbeiten", "Keine Details extrahiert.")
    ws['C29'] = bautagebuch_data.get("Montagegeraete", "Keine Details extrahiert.")
    ws['C31'] = bautagebuch_data.get("Sonstige_Geraeteeinsaetze", "Keine Details extrahiert.")
    ws['C33'] = bautagebuch_data.get("Materialanlieferungen", "Keine Details extrahiert.")
    ws['C35'] = bautagebuch_data.get("Kundenanweisungen", "Keine Details extrahiert.")
    ws['C37'] = bautagebuch_data.get("Informationen_Fremdfirmen_Fremdleistungen", "Keine Details extrahiert.")
    ws['C41'] = bautagebuch_data.get("Maengel_Nachtragsleistungen", "Keine Details extrahiert.")

    # --- Save the filled workbook ---
    try:
        datum_for_filename = bautagebuch_data.get('Datum', config.get_current_date_str()).replace('-', '')
        timestamp_filename = datetime.now().strftime("%Y%m%d_%H%M%S")
        baustelle_name = bautagebuch_data.get("Baustelle", "Unbekannt").replace(" ", "_").replace("/", "-")
        # Sanitize Baustelle name for filename
        valid_baustelle_name = "".join(c for c in baustelle_name if c.isalnum() or c in ('_', '-'))
        if not valid_baustelle_name: valid_baustelle_name = "Allgemein" # Fallback if sanitization results in empty string

        output_filename = f"Bautagebuch_{datum_for_filename}_{valid_baustelle_name}_{timestamp_filename}.xlsx"
        output_file_path = os.path.join(output_dir, output_filename)

        wb.save(output_file_path)
        print(f"\n✅ Bautagebuch Excel guardado en: {output_file_path}")
        return output_file_path
    except Exception as e:
        print(f"❌ Error al guardar el archivo Excel: {e}")
        return None


def save_extended_text_details(bautagebuch_data, current_date_str):
    """Saves extended details to a text file."""
    if not bautagebuch_data:
        print("No hay datos para guardar en archivo de texto.")
        return

    # Usar la clave "Ausgefuehrte_Arbeiten" según el JSON_SCHEMA_FOR_LLM
    # Si el LLM produce otra clave como "ausgefuehrte_arbeiten_details" y quieres esa,
    # deberías ajustar el schema o cómo accedes a este dato aquí.
    extended_details = bautagebuch_data.get("Ausgefuehrte_Arbeiten", "No se proporcionaron detalles.")
    timestamp_filename = datetime.now().strftime("%Y%m%d_%H%M%S")

    extended_txt_filename = os.path.join(
        config.EXTENDED_TASKS_DIR,
        f"bautagebuch_extended_{timestamp_filename}.txt"
    )
    try:
        with open(extended_txt_filename, "w", encoding="utf-8") as f:
            f.write(f"Datum: {bautagebuch_data.get('Datum', current_date_str)}\n") # 'Datum' con mayúscula
            f.write(f"Baustelle: {bautagebuch_data.get('Baustelle', 'N/A')}\n\n") # 'Baustelle' con mayúscula
            f.write("Ausgeführte Arbeiten (Details):\n")
            f.write(extended_details)
        print(f"\n📝 Detalles extendidos guardados en: {extended_txt_filename}")
    except Exception as e:
        print(f"⚠️ Error al guardar el archivo de texto extendido: {e}")